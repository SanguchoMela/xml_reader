import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import xml.etree.ElementTree as ET
# from estilos_excel import dar_estilo_factura

existen = None

# Funcion para procesar las etiquetas de cada archivo xml
def procesar_una_factura(ruta_archivo):
    try:
        # Abrir el archivo XML
        arbol = ET.parse(ruta_archivo)
        # Obtener la raíz del árbol
        ruta = arbol.getroot()
        
        # obtener el comprobante dentro del CDATA
        comprobante_cdata = ruta.find('comprobante').text.strip()

        # parsear el comprobante xml contenido en el CDATA
        comprobante_root = ET.fromstring(comprobante_cdata)
        
        info_tributaria = comprobante_root.find('infoTributaria')
        if info_tributaria is not None:
            razon_social_vendedor = info_tributaria.find('razonSocial').text
            ruc_vendedor = info_tributaria.find('ruc').text
            cod_doc = info_tributaria.find('codDoc').text
            estab = info_tributaria.find('estab').text
            pto_emi = info_tributaria.find('ptoEmi').text
            secuencial = info_tributaria.find('secuencial').text
            
        # # Obtener información de la factura
        factura_info = comprobante_root.find('infoFactura')
        if factura_info is not None:
            fecha_emision_factura = factura_info.find('fechaEmision').text
            razon_social_comprador = factura_info.find('razonSocialComprador').text
            ruc_comprador = factura_info.find('identificacionComprador').text
            total_sin_impuestos = factura_info.find('totalSinImpuestos').text
            importe_total_factura = factura_info.find('importeTotal').text
            propina_elem = factura_info.find('propina')
            if propina_elem is not None:
                propina = propina_elem.text
            else:
                propina = "null"
            
        # Obtener información del total con impuestos
        total_con_impuesto_elem = factura_info.find('totalConImpuestos')
        
        # Inicializar las variables para evitar UnboundLocalError
        subtotal15 = ""
        iva15 = ""
        subtotal0 = ""
        otraTarifa = ""
        otroIVA = ""
        
        for total_impuesto in total_con_impuesto_elem.findall('totalImpuesto'):
            # Extraer los datos del elemento 'totalImpuesto'
            base_imponible = total_impuesto.find('baseImponible').text
            codigo_porcentaje = total_impuesto.find('codigoPorcentaje').text
            valor_iva = total_impuesto.find('valor').text
            
            if codigo_porcentaje == "4":
                subtotal15 = base_imponible
                iva15 = valor_iva
            elif codigo_porcentaje == '0':
                subtotal0 = base_imponible
            else:
                otraTarifa = base_imponible
                otroIVA = valor_iva
                
        # Obtener información de detalles
        detalles_elem = comprobante_root.find('detalles')
        detalles_lista = []
        
        if detalles_elem is not None:
            for detalle in detalles_elem.findall('detalle'):
                
                descripcion = detalle.find('descripcion').text
                cantidad = detalle.find('cantidad').text
                precio_unitario = detalle.find('precioUnitario').text
                descuento = detalle.find('descuento').text
                total_sin_impuesto = detalle.find('precioTotalSinImpuesto').text
                
                # IVA
                iva = ""
                impuestos = detalle.find('impuestos')
                if impuestos is not None:
                    impuesto = impuestos.find('impuesto')
                    if impuesto is not None:
                        iva = impuesto.find('valor').text
                        
                # Detalles adicionales
                adicionales_str = ""
                detalles_adicionales = detalle.find('detallesAdicionales')
                
                if detalles_adicionales is not None:
                    adicionales = []
                    for det in detalles_adicionales.findall('detAdicional'):
                        valor = det.get('valor')
                        adicionales.append(valor)
                    
                    adicionales_str = " | ".join(adicionales)
                    
                # Fila completa
                detalles_lista.append({
                    "Descripcion": descripcion,
                    "Cantidad": cantidad,
                    "Precio Unitario": precio_unitario,
                    "Descuento": descuento,
                    "Total Sin Impuesto": total_sin_impuesto,
                    "IVA": iva,
                    "Adicionales": adicionales_str
                })


        # Agregar información a la lista
        datos_generales = {
            "Razón Social Comprador": razon_social_comprador,
            "RUC Comprador": ruc_comprador,
            "Razón Social Vendedor": razon_social_vendedor,
            "RUC Vendedor": ruc_vendedor,
            "Fecha de Emisión": fecha_emision_factura,
            "Numero de Factura": f"{cod_doc}-{estab}-{pto_emi}-{secuencial}",
            "Propina": propina,
            "Subtotal 0%": subtotal0,
            "Subtotal 15%": subtotal15,
            "Otra tarifa": otraTarifa,
            "Subtotal sin Impuestos": total_sin_impuestos,
            "IVA 15%": iva15,
            "Otro IVA": otroIVA,
            "Total": importe_total_factura,
        }
            
    except ET.ParseError as e:
        print(f"Error en archivo {ruta_archivo}: {e}")
    
    return datos_generales, detalles_lista


def exportar_factura_individual(data, detalles, excel_file):
    try:        
        wb = Workbook()
        ws = wb.active
        ws.title = "Factura"
        
        # Encabezado 
        fila = 1
        for clave, valor in data.items():
            ws.cell(row=fila, column=1, value=clave).font = Font(bold=True)
            ws.cell(row=fila, column=2, value=valor)
            fila +=1 
            
        # Espacio 
        fila += 1
        
        # Título tabla
        ws.cell(row=fila, column=1, value="Detalle").font = Font(bold=True)
        
        fila += 1
        
        # Encabezado tabla
        headers = ["Descripcion", "Cantidad", "Precio Unitario", "Descuento", "Total Sin Impuesto", "IVA", "Adicionales"]
        
        # fila_inicio_tabla = fila
        
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=fila, column=col, value=h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            
        fila += 1
        
        # Filas detalles
        for item in detalles:
            for col, h in enumerate(headers, start=1):
                ws.cell(row=fila, column=col, value=item.get(h, ""))
            fila += 1
            
        # fila_fin_tabla = fila - 1
        # num_filas_tabla = fila_fin_tabla - fila_inicio_tabla + 1
        
        # dar_estilo_factura(excel_file, fila_inicio_tabla, len(headers), num_filas_tabla)
            
        wb.save(excel_file)    
        
        print(f"Factura exportada en '{excel_file}'.")
    except Exception as e:
        print(f"Error al exportar(): {e}")


# ruta_xml = "C:/Users/Hp/Downloads/FacturaInv/Factura.xml"
# excel_salida = "C:/Users/Hp/Downloads/resumen.xlsx"

# datos, detalles = procesar_una_factura(ruta_xml)
# exportar_factura_individual(datos, detalles, excel_salida)