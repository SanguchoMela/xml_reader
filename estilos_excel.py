import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side

def dar_estilo(archivo_excel, df, col_no_centrar):            
    # Agregar estilos con openpyxl
    wb = openpyxl.load_workbook(archivo_excel)
    ws = wb.active
            
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length +2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
        
    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font
        
    col_centrar = [col for col in df.columns if col not in col_no_centrar]
    
    for col_index, column_name in enumerate(df.columns):
        col_letter = chr(65 + col_index) 
        if column_name in col_centrar: 
            for row in range(2, len(df) + 2): 
                ws[f"{col_letter}{row}"].alignment = Alignment(horizontal="center")
    
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row in ws.iter_rows(min_row=2, max_row=len(df) + 1, min_col=1, max_col=len(df.columns)):
        for cell in row:
            cell.border = thin_border
        
    wb.save(archivo_excel)