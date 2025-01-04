import os
import pandas as pd
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side

data_retenciones = []
existen_retenciones = None

def procesar_retenciones(ruta_archivo):
    try:
        # Abrir el archivo XML
        arbol = ET.parse(ruta_archivo)
        # Obtener la raíz del árbol
        ruta = arbol.getroot()
        
        # obtener el comprobante dentro del CDATA
        comprobante_cdata = ruta.find('comprobante').text.strip()
        
        # parsear el comprobante xml contenido en el CDATA
        comprobante_root = ET.fromstring(comprobante_cdata)
        
        info_tributaria = comprobante_root.find('infoTributaria')
        
        if info_tributaria is not None:
            razon_social_agenteRet = info_tributaria.find('razonSocial').text
            ruc_agente = info_tributaria.find('ruc').text
            # cod_doc = info_tributaria.find('codDoc').text
            estab = info_tributaria.find('estab').text
            pto_emi = info_tributaria.find('ptoEmi').text
            secuencial = info_tributaria.find('secuencial').text
            
        # print(f"{razon_social_agenteRet}, {ruc_agente}, Num: {estab}-{pto_emi}-{secuencial}")
            
        # Obtener información de la retencion
        retencion_info = comprobante_root.find('infoCompRetencion')
        if retencion_info is not None:
            fecha_emision_ret = retencion_info.find('fechaEmision').text
            razon_social_sujeto_retenido = retencion_info.find('razonSocialSujetoRetenido').text
            ruc_sujeto_retenido = retencion_info.find('identificacionSujetoRetenido').text
            
        # print(f"{fecha_emision_ret}, {razon_social_sujeto_retenido}, {ruc_sujeto_retenido}")
        
        base_renta = ""
        porcentaje_renta = ""
        valor_renta = ""
        base_iva = ""
        porcentaje_iva = ""
        valor_iva = ""
        num_sustento = ""
        numero_doc_sustento = ""
        numero_doc_sustento_uno = ""
        
        # Información del sustento del documento, con solo un impuesto
        docs_sustento_elem = comprobante_root.find('docsSustento')
        
        if docs_sustento_elem is not None:        
            for doc_sustento in docs_sustento_elem.findall('docSustento'):
                # Extraer datos del elemento docSustento
                numero_doc_sustento_uno = doc_sustento.find('numDocSustento').text
                
                retenciones_elem = doc_sustento.find('retenciones')
                for retencion in retenciones_elem.findall('retencion'):
                    codigo_impuesto_uno = retencion.find('codigo').text
                    base_imponible_uno = retencion.find('baseImponible').text
                    porcentaje_ret_uno = retencion.find('porcentajeRetener').text
                    valor_retenido_uno = retencion.find('valorRetenido').text

                    if codigo_impuesto_uno == "1":
                        base_renta = base_imponible_uno
                        porcentaje_renta = porcentaje_ret_uno
                        valor_renta = valor_retenido_uno
                    elif codigo_impuesto_uno == "2":
                        base_iva = base_imponible_uno
                        porcentaje_iva = porcentaje_ret_uno
                        valor_iva = valor_retenido_uno

                    # print(f"{numero_doc_sustento_uno}")  
                    # print(f"{codigo_impuesto_uno},{base_imponible_uno}, {porcentaje_ret_uno}, {valor_retenido_uno}")     
            
        # Información de sustento del documento, con dos impuestos
        impuestos_elem = comprobante_root.find('impuestos')
        if impuestos_elem is not None:
            for impuesto in impuestos_elem.findall('impuesto'):
                numero_doc_sustento = impuesto.find('numDocSustento').text
                codigo_impuesto = impuesto.find('codigo').text
                base_imponible = impuesto.find('baseImponible').text
                porcentaje_ret = impuesto.find('porcentajeRetener').text
                valor_retenido = impuesto.find('valorRetenido').text
                
                if codigo_impuesto == "1":
                    base_renta = base_imponible
                    porcentaje_renta = porcentaje_ret
                    valor_renta = valor_retenido
                elif codigo_impuesto == "2":
                    base_iva = base_imponible
                    porcentaje_iva = porcentaje_ret
                    valor_iva = valor_retenido
        
                # print(f"{numero_doc_sustento}")  
                # print(f"{codigo_impuesto},{base_imponible}, {porcentaje_ret}, {valor_retenido}")    
                
        if numero_doc_sustento == "":
            num_sustento = numero_doc_sustento_uno
        elif numero_doc_sustento_uno == "":
            num_sustento = numero_doc_sustento

        # print("---------------------------------------")
        
        # Agregar información a la lista
        data_retenciones.append({
            "info_xml": {
                "Razón Social Agente Ret": razon_social_agenteRet,
                "RUC Agente": ruc_agente,
                "Razón Social Sujeto Ret": razon_social_sujeto_retenido,
                "RUC Sujeto": ruc_sujeto_retenido,
                "Fecha de Emisión": fecha_emision_ret,
                "Numero de Retención": f"{estab}-{pto_emi}-{secuencial}",
                "Numero Doc Sustento": num_sustento,
                "Base imponible renta": base_renta,
                "Porcentaje retención renta": porcentaje_renta,
                "Valor retenido renta": valor_renta,
                "Base imponible IVA": base_iva,
                "Porcentaje retención IVA": porcentaje_iva,                
                "Valor retenido IVA": valor_iva                
            }
        })
        
    except ET.ParseError as e:
        print(f"Error en archivo {ruta_archivo}: {e}") 

def exportar_retenciones_excel(data, excel_file):
    try:        
        # Crear DataFrame
        df = pd.DataFrame([fila["info_xml"] for fila in data_retenciones])
        
        # Covertir varias columnas a numeros
        for column in ["Base imponible renta","Porcentaje retención renta", "Valor retenido renta", "Base imponible IVA", "Porcentaje retención IVA", "Valor retenido IVA"]:
            df[column] = pd.to_numeric(df[column], errors='coerce')
            
        df.fillna(0, inplace=True)

        # Escribir DataFrame a Excel
        df.to_excel(excel_file, index=False)
        
        # Agregar estilos con openpyxl
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
                
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length +2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
            
        bold_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold_font
            
        col_no_centrar = ["Razón Social Agente Ret", "Razón Social Sujeto Ret"]
        col_centrar = [col for col in df.columns if col not in col_no_centrar]
        
        for col_index, column_name in enumerate(df.columns):
            col_letter = chr(65 + col_index) 
            if column_name in col_centrar: 
                for row in range(2, len(df) + 2): 
                    ws[f"{col_letter}{row}"].alignment = Alignment(horizontal="center")
        
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row in ws.iter_rows(min_row=2, max_row=len(df) + 1, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.border = thin_border
            
        wb.save(excel_file)
        
        print(f"Datos exportados a '{excel_file}' correctamente.")
    except Exception as e:
        print(f"Error al exportar(): {e}")

def procesar_carpeta_retenciones(ruta_carpeta):
    global existen_retenciones
    existen_retenciones = False 
    
    # listar archivos
    archivos = os.listdir(ruta_carpeta)

    # filtrar solo archivos xml
    archivos_xml = [archivo for archivo in archivos if archivo.endswith(".xml")]
    
    if archivos_xml:
        existen_retenciones = True

    # procesar cada archivo xml
    for archivo_xml in archivos_xml:
        ruta_completa = os.path.join(ruta_carpeta, archivo_xml)
        procesar_retenciones(ruta_completa)
        
    return existen_retenciones

# # ruta de la carpeta con retenciones xml
# carpeta_ret_xml = "C:/Users/Hp/Downloads/drP_ret_nov24"

# # procesar todos los xml en la carpeta
# try:
#     procesar_carpeta_retenciones(carpeta_ret_xml)
# except Exception as e:
#     print(f"Error: {e}")

# # exportar datos a un archivo excel
# excel_ruta = "C:/Users/Hp/Downloads/resumenRet_drP_nov24.xlsx"

# print(exportar_retenciones_excel(data_retenciones, excel_ruta))

# try:
#     exportar_retenciones_excel = (data_retenciones, excel_ruta)
# except Exception as e:
#     print(f"Error al exportar: {e}")